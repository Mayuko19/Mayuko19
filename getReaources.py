import openpyxl as xl

fname = 'ElasticContainerService task patching retirement_July13.xlsx'
wb = xl.load_workbook(fname)
ws = wb.active

for r in ws['A1':'A66']:
    print(ws["r"].value)
